"""
Data processing and Excel generation module
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

import matplotlib
matplotlib.use('Agg')  # Use non-GUI backend
import matplotlib.pyplot as plt
import seaborn as sns
from wordcloud import WordCloud

from config import Config


class DataProcessor:
    """Handles data processing and analysis"""
    
    def __init__(self, config: Config):
        self.config = config
        self.logger = logging.getLogger(__name__)
        sns.set_style('whitegrid')
        
    def create_excel_report(self, products_df: pd.DataFrame, reviews_df: pd.DataFrame):
        """Create comprehensive Excel report with multiple sheets"""
        self.logger.info("Creating Excel report")
        
        excel_path = self.config.REPORTS_DIR / self.config.EXCEL_FILENAME
        
        try:
            # Create Excel writer
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Sheet 1: Product Summary
                self._create_product_summary_sheet(products_df, writer)
                
                # Sheet 2: Specifications Comparison
                self._create_specs_comparison_sheet(products_df, writer)
                
                # Sheet 3: Review Analysis
                self._create_review_analysis_sheet(reviews_df, writer)
                
                # Sheet 4: Brand Analysis
                self._create_brand_analysis_sheet(products_df, writer)
            
            # Apply formatting
            self._apply_excel_formatting(excel_path, products_df, reviews_df)
            
            self.logger.info(f"Excel report created: {excel_path}")
            
        except Exception as e:
            self.logger.error(f"Error creating Excel report: {str(e)}", exc_info=True)
            raise
    
    def _create_product_summary_sheet(self, df: pd.DataFrame, writer):
        """Create product summary sheet with formatting"""
        if df.empty:
            return
        
        # Select and order columns
        summary_cols = ['name', 'price', 'rating', 'num_reviews']
        summary_df = df[summary_cols].copy() if all(col in df.columns for col in summary_cols) else df
        
        # Add calculated fields
        if 'price' in summary_df.columns and 'rating' in summary_df.columns:
            summary_df['value_score'] = summary_df['rating'] / (summary_df['price'] / 100)
        
        summary_df.to_excel(writer, sheet_name='Product Summary', index=False)
    
    def _create_specs_comparison_sheet(self, df: pd.DataFrame, writer):
        """Create specifications comparison matrix"""
        if df.empty or 'specifications' not in df.columns:
            return
        
        try:
            # Extract specifications into separate columns
            specs_data = []
            for idx, row in df.iterrows():
                specs = row.get('specifications', {})
                if isinstance(specs, dict):
                    spec_row = {'product_name': row.get('name', f'Product {idx}')}
                    spec_row.update(specs)
                    specs_data.append(spec_row)
            
            if specs_data:
                specs_df = pd.DataFrame(specs_data)
                specs_df.to_excel(writer, sheet_name='Specifications', index=False)
        
        except Exception as e:
            self.logger.warning(f"Error creating specs comparison: {str(e)}")
    
    def _create_review_analysis_sheet(self, df: pd.DataFrame, writer):
        """Create review analysis sheet"""
        if df.empty:
            return
        
        try:
            # Aggregate review statistics
            analysis_data = []
            
            if 'product_url' in df.columns:
                grouped = df.groupby('product_url')
                
                for url, group in grouped:
                    stats = {
                        'product': url.split('/')[-1][:50],  # Truncate URL
                        'total_reviews': len(group),
                        'avg_rating': group['rating'].mean() if 'rating' in group.columns else 0,
                        'sentiment_score': group['sentiment_score'].mean() if 'sentiment_score' in group.columns else 0,
                        'positive_reviews': len(group[group['sentiment_label'] == 'Positive']) if 'sentiment_label' in group.columns else 0,
                        'negative_reviews': len(group[group['sentiment_label'] == 'Negative']) if 'sentiment_label' in group.columns else 0,
                        'neutral_reviews': len(group[group['sentiment_label'] == 'Neutral']) if 'sentiment_label' in group.columns else 0
                    }
                    analysis_data.append(stats)
            
            if analysis_data:
                analysis_df = pd.DataFrame(analysis_data)
                analysis_df.to_excel(writer, sheet_name='Review Analysis', index=False)
            else:
                # If no grouping possible, write raw data
                review_cols = ['rating', 'title', 'text', 'sentiment_score', 'sentiment_label']
                review_df = df[[col for col in review_cols if col in df.columns]]
                review_df.to_excel(writer, sheet_name='Review Analysis', index=False)
        
        except Exception as e:
            self.logger.warning(f"Error creating review analysis: {str(e)}")
    
    def _create_brand_analysis_sheet(self, df: pd.DataFrame, writer):
        """Create brand analysis pivot table"""
        if df.empty:
            return
        
        try:
            # Extract brand from product name (simple heuristic)
            df_copy = df.copy()
            df_copy['brand'] = df_copy['name'].str.split().str[0]
            
            # Create brand summary
            brand_summary = df_copy.groupby('brand').agg({
                'price': ['mean', 'min', 'max'],
                'rating': 'mean',
                'num_reviews': 'sum',
                'name': 'count'
            }).round(2)
            
            brand_summary.columns = ['_'.join(col).strip() for col in brand_summary.columns.values]
            brand_summary.reset_index(inplace=True)
            
            brand_summary.to_excel(writer, sheet_name='Brand Analysis', index=False)
        
        except Exception as e:
            self.logger.warning(f"Error creating brand analysis: {str(e)}")
    
    def _apply_excel_formatting(self, excel_path: Path, products_df: pd.DataFrame, reviews_df: pd.DataFrame):
        """Apply conditional formatting and styling to Excel file"""
        try:
            wb = load_workbook(excel_path)
            
            # Format Product Summary sheet
            if 'Product Summary' in wb.sheetnames:
                ws = wb['Product Summary']
                self._format_product_summary(ws, products_df)
            
            # Format Review Analysis sheet
            if 'Review Analysis' in wb.sheetnames:
                ws = wb['Review Analysis']
                self._format_review_analysis(ws)
            
            wb.save(excel_path)
            
        except Exception as e:
            self.logger.warning(f"Error applying Excel formatting: {str(e)}")
    
    def _format_product_summary(self, ws, df):
        """Apply formatting to product summary sheet"""
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Price conditional formatting (if price column exists)
        try:
            price_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value and 'price' in str(cell.value).lower():
                    price_col = idx
                    break
            
            if price_col:
                ws.conditional_formatting.add(
                    f'{ws.cell(2, price_col).column_letter}2:{ws.cell(ws.max_row, price_col).column_letter}{ws.max_row}',
                    ColorScaleRule(
                        start_type='min', start_color='63BE7B',
                        mid_type='percentile', mid_value=50, mid_color='FFEB84',
                        end_type='max', end_color='F8696B'
                    )
                )
        except Exception as e:
            self.logger.warning(f"Error applying price formatting: {str(e)}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _format_review_analysis(self, ws):
        """Apply formatting to review analysis sheet"""
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_visualizations(self, products_df: pd.DataFrame, reviews_df: pd.DataFrame):
        """Create data visualizations"""
        self.logger.info("Creating visualizations")
        
        viz_dir = self.config.REPORTS_DIR / 'visualizations'
        viz_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            # Price distribution
            if not products_df.empty and 'price' in products_df.columns:
                self._plot_price_distribution(products_df, viz_dir)
            
            # Rating distribution
            if not products_df.empty and 'rating' in products_df.columns:
                self._plot_rating_distribution(products_df, viz_dir)
            
            # Brand comparison
            if not products_df.empty and 'name' in products_df.columns:
                self._plot_brand_comparison(products_df, viz_dir)
            
            # Sentiment analysis
            if not reviews_df.empty and 'sentiment_label' in reviews_df.columns:
                self._plot_sentiment_analysis(reviews_df, viz_dir)
            
            # Word cloud
            if not reviews_df.empty and 'text' in reviews_df.columns:
                self._create_word_cloud(reviews_df, viz_dir)
            
            self.logger.info(f"Visualizations saved to {viz_dir}")
            
        except Exception as e:
            self.logger.error(f"Error creating visualizations: {str(e)}", exc_info=True)
    
    def _plot_price_distribution(self, df: pd.DataFrame, viz_dir: Path):
        """Create price distribution histogram"""
        plt.figure(figsize=(10, 6))
        plt.hist(df['price'].dropna(), bins=20, color='skyblue', edgecolor='black')
        plt.xlabel('Price ($)')
        plt.ylabel('Frequency')
        plt.title('Product Price Distribution')
        plt.grid(axis='y', alpha=0.75)
        plt.tight_layout()
        plt.savefig(viz_dir / 'price_distribution.png', dpi=300, bbox_inches='tight')
        plt.close()
    
    def _plot_rating_distribution(self, df: pd.DataFrame, viz_dir: Path):
        """Create rating distribution bar chart"""
        plt.figure(figsize=(10, 6))
        rating_counts = df['rating'].value_counts().sort_index()
        rating_counts.plot(kind='bar', color='coral', edgecolor='black')
        plt.xlabel('Rating')
        plt.ylabel('Number of Products')
        plt.title('Product Rating Distribution')
        plt.xticks(rotation=0)
        plt.tight_layout()
        plt.savefig(viz_dir / 'rating_distribution.png', dpi=300, bbox_inches='tight')
        plt.close()
    
    def _plot_brand_comparison(self, df: pd.DataFrame, viz_dir: Path):
        """Create brand comparison chart"""
        try:
            df_copy = df.copy()
            df_copy['brand'] = df_copy['name'].str.split().str[0]
            
            brand_stats = df_copy.groupby('brand').agg({
                'price': 'mean',
                'rating': 'mean',
                'name': 'count'
            }).sort_values('name', ascending=False).head(10)
            
            fig, axes = plt.subplots(1, 2, figsize=(15, 6))
            
            # Average price by brand
            brand_stats['price'].plot(kind='barh', ax=axes[0], color='steelblue')
            axes[0].set_xlabel('Average Price ($)')
            axes[0].set_ylabel('Brand')
            axes[0].set_title('Average Price by Brand')
            
            # Average rating by brand
            brand_stats['rating'].plot(kind='barh', ax=axes[1], color='seagreen')
            axes[1].set_xlabel('Average Rating')
            axes[1].set_ylabel('Brand')
            axes[1].set_title('Average Rating by Brand')
            
            plt.tight_layout()
            plt.savefig(viz_dir / 'brand_comparison.png', dpi=300, bbox_inches='tight')
            plt.close()
            
        except Exception as e:
            self.logger.warning(f"Error creating brand comparison: {str(e)}")
    
    def _plot_sentiment_analysis(self, df: pd.DataFrame, viz_dir: Path):
        """Create sentiment analysis visualizations"""
        try:
            fig, axes = plt.subplots(1, 2, figsize=(15, 6))
            
            # Sentiment distribution pie chart
            sentiment_counts = df['sentiment_label'].value_counts()
            colors = {'Positive': '#90EE90', 'Neutral': '#FFD700', 'Negative': '#FF6B6B'}
            sentiment_colors = [colors.get(label, '#CCCCCC') for label in sentiment_counts.index]
            
            axes[0].pie(sentiment_counts.values, labels=sentiment_counts.index, 
                       autopct='%1.1f%%', colors=sentiment_colors, startangle=90)
            axes[0].set_title('Sentiment Distribution')
            
            # Sentiment score distribution
            if 'sentiment_score' in df.columns:
                axes[1].hist(df['sentiment_score'].dropna(), bins=30, 
                            color='mediumpurple', edgecolor='black')
                axes[1].set_xlabel('Sentiment Score')
                axes[1].set_ylabel('Frequency')
                axes[1].set_title('Sentiment Score Distribution')
                axes[1].axvline(x=0, color='red', linestyle='--', label='Neutral')
                axes[1].legend()
            
            plt.tight_layout()
            plt.savefig(viz_dir / 'sentiment_analysis.png', dpi=300, bbox_inches='tight')
            plt.close()
            
        except Exception as e:
            self.logger.warning(f"Error creating sentiment visualization: {str(e)}")
    
    def _create_word_cloud(self, df: pd.DataFrame, viz_dir: Path):
        """Create word cloud from review text"""
        try:
            # Combine all review text
            text = ' '.join(df['text'].dropna().astype(str))
            
            if len(text) > 0:
                wordcloud = WordCloud(width=800, height=400, 
                                     background_color='white',
                                     colormap='viridis',
                                     max_words=100).generate(text)
                
                plt.figure(figsize=(12, 6))
                plt.imshow(wordcloud, interpolation='bilinear')
                plt.axis('off')
                plt.title('Most Common Words in Reviews', fontsize=16, pad=20)
                plt.tight_layout()
                plt.savefig(viz_dir / 'word_cloud.png', dpi=300, bbox_inches='tight')
                plt.close()
                
        except Exception as e:
            self.logger.warning(f"Error creating word cloud: {str(e)}")